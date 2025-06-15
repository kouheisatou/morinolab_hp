#!/usr/bin/env python3
"""
Generate CSV files and thumbnail images from a single Excel workbook in which
• each worksheet corresponds to a data model (Theme, Lecture, Member …)
• the first row holds the property names (exactly matching @/models fields)
• the "thumbnail" column contains an *embedded* image (picture) instead of a file path

For every worksheet <SheetName>, the script produces:
    public/<sheet-lower>.csv          – UTF-8 CSV with the worksheet data (+thumbnail path)
    public/img/<sheet-lower>/<ID>.jpg – JPEG thumbnail extracted & resized

Usage example
-------------
python scripts/generate_excel_assets.py \
    --excel-path data/content.xlsx

Dependencies
------------
    pip install openpyxl openpyxl-image-loader pillow pandas

The script is designed for GitHub Actions but works locally as well.
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Any, Dict, List, Tuple, Iterator
import hashlib
import io
import zipfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl_image_loader import SheetImageLoader  # type: ignore
from PIL import Image
from openpyxl.utils import get_column_letter

# Thumbnail target size (max width, max height)
THUMB_SIZE: Tuple[int, int] = (640, 360)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Generate CSV & thumbnails from workbook")
    p.add_argument("--excel-path", required=True, help="Path to the workbook (.xlsx)")
    p.add_argument("--csv-dir", default="public/generated_contents", help="Directory root for CSV files")
    p.add_argument("--img-dir", default="public/generated_contents/img", help="Directory root for thumbnail images")
    p.add_argument("--encoding", default="utf-8", help="CSV encoding")
    return p.parse_args()


def ensure_dir(path: Path) -> None:
    if not path.exists():
        path.mkdir(parents=True, exist_ok=True)


def save_thumbnail(image: Image.Image, dest: Path) -> None:
    # Resize proportionally so that neither width nor height exceeds THUMB_SIZE
    image.thumbnail(THUMB_SIZE, Image.LANCZOS)
    if image.mode in ("RGBA", "P"):
        image = image.convert("RGB")
    ensure_dir(dest.parent)
    image.save(dest, format="JPEG", quality=85, optimize=True)


# ---------------------------------------------------------------------------
# Core logic per sheet
# ---------------------------------------------------------------------------

def process_sheet(ws: Worksheet, root_dir: Path, excel_dir: Path, encoding: str, img_iter: Iterator[bytes]) -> None:
    sheet_name = ws.title.strip()
    sheet_key = sheet_name.lower()
    sheet_dir = root_dir / sheet_key
    ensure_dir(sheet_dir)

    # Use pandas to read sheet values quickly (values only)
    df = pd.DataFrame(ws.values)
    if df.empty:
        print(f"[INFO] Sheet '{sheet_name}' is empty – skipped")
        return

    # First row = headers
    headers = [str(h).strip() if h is not None else "" for h in df.iloc[0]]
    df = df.iloc[1:].reset_index(drop=True)
    df.columns = headers

    if "id" not in (h.lower() for h in headers):
        print(f"[WARN] Sheet '{sheet_name}' does not contain 'id' column – skipped")
        return

    # Build mapping (row_index -> PIL.Image) for embedded images
    image_loader: SheetImageLoader | None = None
    try:
        image_loader = SheetImageLoader(ws)
    except Exception as exc:  # pragma: no cover
        # No images or unsupported (rare) – we continue without thumbnails
        print(f"[INFO] No embedded images accessible for sheet '{sheet_name}': {exc}")

    # Determine column letter(s) for 'thumbnail' (case-insensitive)
    thumb_col_idx = next((i for i, h in enumerate(headers) if str(h).lower() == "thumbnail"), None)

    records: List[Dict[str, Any]] = []
    for idx, row in df.iterrows():
        rec: Dict[str, Any] = {h: row[h] for h in headers}

        rec_id_raw = rec.get("id") or rec.get("ID") or rec.get("Id")
        rec_id = str(rec_id_raw).strip() if rec_id_raw is not None else ""
        if not rec_id:
            # skip empty rows
            continue

        # Thumbnail extraction – first try embedded picture, then fallback to file path in the cell
        if thumb_col_idx is not None:
            thumb_header = headers[thumb_col_idx]

            saved_thumb = False

            # 1) Embedded image case -------------------------------------
            if image_loader is not None:
                col_letter = get_column_letter(thumb_col_idx + 1)
                cell_addr = f"{col_letter}{idx + 2}"  # Data starts at row 2
                if image_loader.image_in(cell_addr):
                    pil_img = image_loader.get(cell_addr)

                    # Compute hash filename
                    buf = io.BytesIO()
                    pil_img.save(buf, format="PNG")
                    img_hash = hashlib.md5(buf.getvalue()).hexdigest()

                    dest_img = sheet_dir / f"{img_hash}.jpg"
                    save_thumbnail(pil_img, dest_img)

                    rec[thumb_header] = dest_img.name  # store filename only
                    saved_thumb = True

            # 2) Fallback: path string in cell ---------------------------
            if not saved_thumb:
                path_value = str(rec.get(thumb_header, "")).strip()
                if path_value:
                    src_img_path = (excel_dir / path_value).resolve() if not Path(path_value).is_absolute() else Path(path_value)
                    if src_img_path.exists():
                        try:
                            from PIL import Image as PILImage  # local import to avoid circular
                            pil_img = PILImage.open(src_img_path)

                            # Compute hash filename based on file bytes
                            with open(src_img_path, "rb") as fh:
                                img_hash = hashlib.md5(fh.read()).hexdigest()

                            dest_img = sheet_dir / f"{img_hash}.jpg"
                            save_thumbnail(pil_img, dest_img)

                            rec[thumb_header] = dest_img.name
                            saved_thumb = True
                        except Exception as exc:  # pragma: no cover
                            print(f"[WARN] Failed to process thumbnail '{src_img_path}': {exc}")

            # 3) No thumbnail extracted ----------------------------------
            if not saved_thumb:
                rec[thumb_header] = ""

            # 4) Last-resort: assign next media image if available ------
            if not saved_thumb:
                try:
                    img_bytes = next(img_iter)
                    img_hash = hashlib.md5(img_bytes).hexdigest()
                    dest_img = sheet_dir / f"{img_hash}.jpg"
                    from PIL import Image as PILImage
                    pil_img = PILImage.open(io.BytesIO(img_bytes))
                    save_thumbnail(pil_img, dest_img)
                    rec[thumb_header] = dest_img.name
                    saved_thumb = True
                except StopIteration:
                    pass  # No more fallback images

        records.append(rec)

    if not records:
        print(f"[INFO] No valid rows found in sheet '{sheet_name}' – skipped")
        return

    # Write CSV
    csv_path = sheet_dir / f"{sheet_key}.csv"
    ensure_dir(csv_path.parent)
    with open(csv_path, "w", newline="", encoding=encoding) as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        writer.writerows(records)
    print(f"[INFO] {sheet_name}: wrote {len(records)} records to {csv_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    args = parse_args()

    wb_path = Path(args.excel_path)
    if not wb_path.exists():
        sys.exit(f"Excel workbook not found: {wb_path}")

    root_dir = Path(args.csv_dir)
    ensure_dir(root_dir)

    wb = load_workbook(wb_path, data_only=True)

    # Gather any images present in the workbook archive (xl/media/*)
    with zipfile.ZipFile(wb_path) as zf:
        media_files = [name for name in zf.namelist() if name.startswith("xl/media/")]
        media_bytes_list = [zf.read(name) for name in media_files]

    media_iter = iter(media_bytes_list)

    for ws in wb.worksheets:
        process_sheet(ws, root_dir, wb_path.parent, args.encoding, media_iter)

    print("[INFO] All sheets processed.")


if __name__ == "__main__":
    main() 