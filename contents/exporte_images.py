from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from pathlib import Path
import shutil

# 1. 解凍済み画像のパス（例: xl/media/image1.png ...）
media_dir = Path("unzipped_excel/xl/media")
media_files = sorted(media_dir.glob("image*.*"))

# 2. Excelブック読み込み
wb = load_workbook("contents_db.xlsx")
cell_to_sheet_map = {}

# 3. 各セルにある画像を収集
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    loader = SheetImageLoader(sheet)
    for row in sheet.iter_rows():
        for cell in row:
            coord = cell.coordinate
            if loader.image_in(coord):
                cell_to_sheet_map[f"{sheetname}!{coord}"] = {
                    "sheet": sheetname,
                    "cell": coord
                }

# 4. セル順に画像を対応させて保存
output_dir = Path("exported_cell_images")
output_dir.mkdir(exist_ok=True)
for (cell_ref, meta), media_file in zip(cell_to_sheet_map.items(), media_files):
    ext = media_file.suffix
    filename = f"{meta['sheet']}_{meta['cell']}{ext}"
    shutil.copy(media_file, output_dir / filename)
    print(f"✓ {cell_ref} → {filename}")